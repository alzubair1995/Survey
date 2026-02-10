from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from sqlalchemy import func
from datetime import datetime, time
from io import BytesIO

from extensions import db
from utils import roles_required
from models.response import SurveyResponse

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


from io import BytesIO
from pathlib import Path

from flask import send_file, request, current_app
from sqlalchemy import func

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib import colors

import arabic_reshaper
from bidi.algorithm import get_display
from flask import redirect, url_for





admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

@admin_bp.route("/dashboard")
@login_required
@roles_required("admin")
def dashboard():
    total = db.session.query(func.count(SurveyResponse.id)).scalar() or 0

    # Distributions (counts)
    gender_counts = dict(
        db.session.query(SurveyResponse.gender, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.gender).all()
    )
    stage_counts = dict(
        db.session.query(SurveyResponse.education_stage, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.education_stage).all()
    )
    satisfaction_counts = dict(
        db.session.query(SurveyResponse.satisfaction, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.satisfaction).all()
    )
    understanding_counts = dict(
        db.session.query(SurveyResponse.understanding_help, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.understanding_help).all()
    )
    device_counts = dict(
        db.session.query(SurveyResponse.device, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.device).all()
    )
    internet_counts = dict(
        db.session.query(SurveyResponse.internet_quality, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.internet_quality).all()
    )
    platform_counts = dict(
        db.session.query(SurveyResponse.platform_ease, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.platform_ease).all()
    )
    interaction_counts = dict(
        db.session.query(SurveyResponse.teacher_interaction, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.teacher_interaction).all()
    )
    preference_counts = dict(
        db.session.query(SurveyResponse.study_preference, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.study_preference).all()
    )
    continue_counts = dict(
        db.session.query(SurveyResponse.continue_elearning, func.count(SurveyResponse.id))
        .group_by(SurveyResponse.continue_elearning).all()
    )

    # Last 7 days trend (Baghdad date)
    # We use SQLite date() + timezone shift for better matching
    baghdad_day = func.date(func.datetime(SurveyResponse.created_at, "+3 hours"))
    last_days = db.session.query(baghdad_day, func.count(SurveyResponse.id))\
        .group_by(baghdad_day)\
        .order_by(baghdad_day.desc())\
        .limit(7).all()

    # reverse to show oldest->newest
    last_days = list(reversed(last_days))
    trend_labels = [d for d, _ in last_days]
    trend_values = [c for _, c in last_days]

    # Latest items
    latest = SurveyResponse.query.order_by(SurveyResponse.created_at.desc()).limit(10).all()

    # KPIs (simple “top” values)
    def top3(d):
        return sorted(d.items(), key=lambda x: x[1], reverse=True)[:3]

    kpis = {
        "top_device": top3(device_counts),
        "top_stage": top3(stage_counts),
        "top_preference": top3(preference_counts),
        "top_satisfaction": top3(satisfaction_counts),
    }

    return render_template(
        "admin_dashboard.html",
        total=total,
        gender_counts=gender_counts,
        stage_counts=stage_counts,
        satisfaction_counts=satisfaction_counts,
        understanding_counts=understanding_counts,
        device_counts=device_counts,
        internet_counts=internet_counts,
        platform_counts=platform_counts,
        interaction_counts=interaction_counts,
        preference_counts=preference_counts,
        continue_counts=continue_counts,
        trend_labels=trend_labels,
        trend_values=trend_values,
        latest=latest,
        kpis=kpis,
    )

def _parse_range(f, t):
    df = datetime.strptime(f, "%Y-%m-%d").date()
    dt = datetime.strptime(t, "%Y-%m-%d").date()
    return datetime.combine(df, time.min), datetime.combine(dt, time.max)

@admin_bp.route("/export/excel")
@login_required
@roles_required("admin")
def export_excel():
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")

    start_dt, end_dt = _parse_range(date_from, date_to)

    items = (
        SurveyResponse.query
        .filter(SurveyResponse.created_at >= start_dt, SurveyResponse.created_at <= end_dt)
        .order_by(SurveyResponse.created_at.asc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Survey Responses"

    title = f"Survey Report ({date_from} to {date_to})"
    ws.merge_cells("A1:L12")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
    "ID",
    "الجنس",
    "المرحلة الدراسية",
    "الرضا عن التعلم الإلكتروني",
    "هل يساعدك على فهم المادة؟",
    "الجهاز المستخدم",
    "جودة الانترنت",
    "سهولة المنصة",
    "التفاعل مع المدرس",
    "تفضيل الدراسة",
    "الاستمرار بالتعلم الإلكتروني",
    "التاريخ",
]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="EEF2FF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r in items:
        ws.append([
        r.id,
        r.gender,
        r.education_stage,
        r.satisfaction,
        r.understanding_help,
        r.device,
        r.internet_quality,
        r.platform_ease,
        r.teacher_interaction,
        r.study_preference,
        r.continue_elearning,
        r.created_at.strftime("%Y-%m-%d %H:%M"),
    ])


    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [6, 10, 14, 20, 20, 14, 14, 14, 16, 14, 18, 18]

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"survey_report_{date_from}_to_{date_to}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
def _ar(text: str) -> str:
    if text is None:
        text = ""
    return get_display(arabic_reshaper.reshape(str(text)))
@admin_bp.route("/export/pdf")
@login_required
@roles_required("admin")
def export_pdf():
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")

    if not date_from or not date_to:
        # يرجع المستخدم للداش بورد بدون كسر
        return redirect(url_for("admin.dashboard"))

    # نفس منطق Excel (تاريخ بغداد)
    baghdad_date = func.date(func.datetime(SurveyResponse.created_at, "+3 hours"))

    items = (
        SurveyResponse.query
        .filter(baghdad_date >= date_from, baghdad_date <= date_to)
        .order_by(SurveyResponse.created_at.asc())
        .all()
    )

    def count_by(col):
        return dict(
            db.session.query(col, func.count(SurveyResponse.id))
            .filter(baghdad_date >= date_from, baghdad_date <= date_to)
            .group_by(col)
            .all()
        )

    # إحصائيات
    gender = count_by(SurveyResponse.gender)
    stage = count_by(SurveyResponse.education_stage)
    satisfaction = count_by(SurveyResponse.satisfaction)
    understanding = count_by(SurveyResponse.understanding_help)
    device = count_by(SurveyResponse.device)
    internet = count_by(SurveyResponse.internet_quality)
    platform = count_by(SurveyResponse.platform_ease)
    interaction = count_by(SurveyResponse.teacher_interaction)
    preference = count_by(SurveyResponse.study_preference)
    cont = count_by(SurveyResponse.continue_elearning)

    # ✅ Register Tajawal font (embedded in PDF)
    font_path = Path(current_app.root_path) / "static" / "fonts" / "Tajawal-Regular.ttf"
    if font_path.exists():
        pdfmetrics.registerFont(TTFont("Tajawal", str(font_path)))
        base_font = "Tajawal"
    else:
        base_font = "Helvetica"  # fallback

    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    w, h = A4

    def draw_title(txt, y):
        c.setFont(base_font, 18)
        c.setFillColor(colors.HexColor("#0F172A"))
        c.drawRightString(w - 2*cm, y, _ar(txt))

    def draw_sub(txt, y):
        c.setFont(base_font, 12)
        c.setFillColor(colors.HexColor("#475569"))
        c.drawRightString(w - 2*cm, y, _ar(txt))

    def draw_section(txt, y):
        c.setFont(base_font, 14)
        c.setFillColor(colors.HexColor("#0F172A"))
        c.drawRightString(w - 2*cm, y, _ar(txt))
        c.setStrokeColor(colors.HexColor("#CBD5E1"))
        c.line(2*cm, y-6, w-2*cm, y-6)

    def draw_kv(dic, y, max_rows=10):
        c.setFont(base_font, 12)
        c.setFillColor(colors.HexColor("#0F172A"))
        items_sorted = sorted(dic.items(), key=lambda x: x[1], reverse=True)
        if not items_sorted:
            c.drawRightString(w - 2*cm, y, _ar("لا توجد بيانات"))
            return y - 18

        for i, (k, v) in enumerate(items_sorted[:max_rows]):
            c.drawRightString(w - 2*cm, y - i*16, _ar(f"{k} : {v}"))
        return y - min(len(items_sorted), max_rows)*16 - 8

    # Header
    draw_title("تقرير إحصائيات الاستبيان", h - 2.2*cm)
    draw_sub(f"الفترة: {date_from}  إلى  {date_to}", h - 3.1*cm)
    draw_sub(f"عدد المشاركات: {len(items)}", h - 3.8*cm)

    y = h - 5.0*cm

    # Sections (multiple pages)
    blocks = [
        ("الجنس", gender),
        ("المرحلة الدراسية", stage),
        ("الرضا عن التعلم الإلكتروني", satisfaction),
        ("هل يساعد على فهم المادة؟", understanding),
        ("الجهاز المستخدم", device),
        ("جودة الإنترنت", internet),
        ("سهولة المنصة", platform),
        ("التفاعل مع المدرس", interaction),
        ("تفضيل طريقة الدراسة", preference),
        ("الاستمرار بالتعلم الإلكتروني", cont),
    ]

    for label, data in blocks:
        if y < 5.2*cm:
            c.showPage()
            y = h - 3*cm

        draw_section(label, y)
        y = draw_kv(data, y - 22)

    # Footer
    c.setFont(base_font, 10)
    c.setFillColor(colors.HexColor("#64748B"))
    c.drawString(2*cm, 1.6*cm, _ar("تم إنشاء هذا التقرير تلقائيًا من نظام الاستبيان."))

    c.save()
    bio.seek(0)

    filename = f"survey_report_{date_from}_to_{date_to}.pdf"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )
